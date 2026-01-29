import os
import sys
import django
from pathlib import Path

# Setup Django
BASE_DIR = Path(__file__).resolve().parent
sys.path.append(str(BASE_DIR))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'backend.settings')
django.setup()

from core.models import OpenCourtApplication, User
import openpyxl
from datetime import datetime
from django.utils. dateparse import parse_date

def load_excel_data(file_path):
    """Load data from Excel file"""
    
    print(f"📂 Loading Excel file: {file_path}")
    
    if not os.path.exists(file_path):
        print(f"❌ File not found:  {file_path}")
        return
    
    try:
        # Load workbook
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        
        print(f"📊 Worksheet: {sheet.title}")
        print(f"📏 Total rows: {sheet.max_row}")
        
        # Get or create a default user for created_by
        default_user, _ = User.objects.get_or_create(
            username='system',
            defaults={
                'role': 'ADMIN',
                'first_name': 'System',
                'last_name': 'User'
            }
        )
        
        created_count = 0
        updated_count = 0
        error_count = 0
        errors = []
        
        # Process each row (skip header)
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # Extract data from columns
                sr_no = row[0]  # Sr.No
                dairy_no = row[1]  # Dairy No
                name = row[2]  # Name
                contact = row[3]  # Contact
                marked_to = row[4]  # Marked To
                date_value = row[5]  # Date
                marked_by = row[6]  # Marked By
                timeline = row[7]  # Timeline
                police_station = row[8]  # P.S
                division = row[9]  # DIVISION
                category = row[10]  # Category
                status_value = row[11]  # Status
                days = row[12]  # Days
                feedback_value = row[13]  # Feedback
                dairy_ps = row[14] if len(row) > 14 else ''  # Dairy PS
                
                # Skip if no serial number
                if not sr_no:
                    continue
                
                # Parse date
                if isinstance(date_value, datetime):
                    date_value = date_value.date()
                elif isinstance(date_value, str) and date_value: 
                    try:
                        date_value = parse_date(date_value)
                    except:
                        date_value = None
                else:
                    date_value = None
                
                # Parse days
                if days: 
                    try:
                        days = int(days)
                    except: 
                        days = None
                else:
                    days = None
                
                # Map status
                status_mapping = {
                    'PENDING': 'PENDING',
                    'HEARD': 'HEARD',
                    'REFERRED': 'REFERRED',
                    'CLOSED': 'CLOSED',
                    'Pending': 'PENDING',
                    'Heard': 'HEARD',
                    'Referred': 'REFERRED',
                    'Closed': 'CLOSED',
                }
                mapped_status = status_mapping.get(str(status_value).strip(), 'PENDING') if status_value else 'PENDING'
                
                # Map feedback
                feedback_mapping = {
                    'POSITIVE': 'POSITIVE',
                    'NEGATIVE': 'NEGATIVE',
                    'PENDING': 'PENDING',
                    'Positive': 'POSITIVE',
                    'Negative':  'NEGATIVE',
                    'Pending': 'PENDING',
                }
                mapped_feedback = feedback_mapping.get(str(feedback_value).strip(), 'PENDING') if feedback_value else 'PENDING'
                
                # Prepare application data
                application_data = {
                    'dairy_no': str(dairy_no) if dairy_no else '',
                    'name': str(name) if name else '',
                    'contact': str(contact) if contact else '',
                    'marked_to': str(marked_to) if marked_to else '',
                    'date': date_value,
                    'marked_by': str(marked_by) if marked_by else '',
                    'timeline': str(timeline) if timeline else '',
                    'police_station': str(police_station) if police_station else '',
                    'division': str(division) if division else '',
                    'category': str(category) if category else '',
                    'status': mapped_status,
                    'days': days,
                    'feedback': mapped_feedback,
                    'dairy_ps': str(dairy_ps) if dairy_ps else '',
                    'created_by': default_user,
                }
                
                # Create or update application
                application, created = OpenCourtApplication.objects.update_or_create(
                    sr_no=int(sr_no),
                    defaults=application_data
                )
                
                if created:
                    created_count += 1
                    print(f"✅ Created: Row {row_num} - {name}")
                else: 
                    updated_count += 1
                    print(f"🔄 Updated: Row {row_num} - {name}")
                    
            except Exception as e:
                error_count += 1
                error_msg = f"Row {row_num}: {str(e)}"
                errors.append(error_msg)
                print(f"❌ Error: {error_msg}")
        
        # Summary
        print("\n" + "="*50)
        print("📊 IMPORT SUMMARY")
        print("="*50)
        print(f"✅ Created:  {created_count}")
        print(f"🔄 Updated: {updated_count}")
        print(f"❌ Errors: {error_count}")
        print(f"📝 Total Processed: {created_count + updated_count}")
        
        if errors:
            print("\n⚠️ ERRORS:")
            for error in errors[: 10]:  # Show first 10 errors
                print(f"  - {error}")
            if len(errors) > 10:
                print(f"  ... and {len(errors) - 10} more errors")
        
        print("="*50)
        
    except Exception as e:
        print(f"❌ Fatal error: {str(e)}")
        import traceback
        traceback.print_exc()


if __name__ == '__main__':
    # Your Excel file path
    excel_file_path = r"D:\\seven semester\\FYP\\daily open court\\April3File.xlsx"
    load_excel_data(excel_file_path)